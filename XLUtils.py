import openpyxl #(import Library openpyxl)

def getRowCount(File,sheetName):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumncount(File,sheetName):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(sheetName)
    return(sheet.max_column)

def readData(File,sheetName,rownum,columnno):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum, column=columnno).value

def writeData(File,sheetName,rownum,columnno,data):
    workbook = openpyxl.load_workbook(File)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=rownum, column=columnno).value = data
    workbook.save(File)
